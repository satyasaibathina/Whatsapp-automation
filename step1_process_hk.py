"""
STEP 1 — Housekeeping Payout Processor
Filters HK Excel: Status=Pay, RM Status=Pending
Saves a styled Excel with merged OM Name cells
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
import os

SHEET_NAME = "System Format"


def process_hk_file(input_path, output_folder="output"):
    print(f"   📊 Loading HK: {input_path}")
    df = pd.read_excel(input_path, sheet_name=SHEET_NAME)

    filtered_df = df[
        (df["Status"].str.strip().str.lower() == "pay") &
        (df["RM Status"].str.strip().str.lower() == "pending") &
        (df["Site Code"].notna()) &
        (df["Site Code"].astype(str).str.strip() != "") &
        (df["OM Name"].notna()) &
        (df["OM Name"].astype(str).str.strip() != "")
    ]

    pivot_df = (
        filtered_df[["OM Name", "Site Code", "OM Status", "RM Status", "Status"]]
        .drop_duplicates(subset=["OM Name", "Site Code"])
        .sort_values(by=["OM Name", "Site Code"])
        .reset_index(drop=True)
    )

    total = len(pivot_df)
    print(f"   Filtered records: {total}")

    if total == 0:
        print("   ✅ No HK pending records.")
        return None, 0

    os.makedirs(output_folder, exist_ok=True)
    output_file = os.path.join(output_folder, "HK_Pending.xlsx")
    pivot_df.to_excel(output_file, sheet_name="Pivot", index=False)

    wb = load_workbook(output_file)
    ws = wb["Pivot"]

    header_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
    thin        = Side(style="thin")
    border      = Border(left=thin, right=thin, top=thin, bottom=thin)
    om_fill1    = PatternFill(start_color="E6F4EA", end_color="E6F4EA", fill_type="solid")
    om_fill2    = PatternFill(start_color="DFF1DD", end_color="DFF1DD", fill_type="solid")

    for cell in ws[1]:
        cell.fill      = header_fill
        cell.font      = Font(color="000000", bold=True)
        cell.border    = border
        cell.alignment = Alignment(horizontal="center", vertical="center")

    row = 2
    toggle = True
    while row <= ws.max_row:
        om = ws[f"A{row}"].value
        if om:
            start = row
            while row + 1 <= ws.max_row and ws[f"A{row+1}"].value == om:
                row += 1
            end = row
            if end > start:
                ws.merge_cells(start_row=start, start_column=1, end_row=end, end_column=1)
            fill = om_fill1 if toggle else om_fill2
            toggle = not toggle
            for r in range(start, end + 1):
                for c in range(1, 6):
                    cell           = ws.cell(row=r, column=c)
                    cell.fill      = fill
                    cell.border    = border
                    cell.alignment = Alignment(horizontal="center", vertical="center")
        row += 1

    for col in ws.columns:
        max_len = max((len(str(c.value)) for c in col if c.value), default=10)
        ws.column_dimensions[col[0].column_letter].width = max_len + 4

    wb.save(output_file)
    print(f"   ✅ Saved: {output_file}")
    return output_file, total
