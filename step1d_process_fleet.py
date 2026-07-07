"""
IC Fleet Monthly Payout Processor
===================================
Exact logic:
  - Sheet: "System Format"
  - Filter: Payout Status = "pay"
  - Filter: Site Status dynamically based on date (before 10th: committed/not committed; after 10th: also approved/confirmed)
  - Exclude blank: Site Code, OM Name, Client Name
  - Output columns: OM Name, Site Code, Site Status, Payout Status, Client Name
  - Style: merged OM Name cells, alternating green rows
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
import os

SHEET_NAME = "System Format"


def save_and_style_fleet(pivot_df, output_file):
    pivot_df.to_excel(output_file, sheet_name="FLEET", index=False)

    # ── Style ────────────────────────────────────────────
    wb = load_workbook(output_file)
    ws = wb["FLEET"]

    header_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
    thin        = Side(style="thin")
    border      = Border(left=thin, right=thin, top=thin, bottom=thin)
    om_fill1    = PatternFill(start_color="E6F4EA", end_color="E6F4EA", fill_type="solid")
    om_fill2    = PatternFill(start_color="DFF1DD", end_color="DFF1DD", fill_type="solid")

    # Header
    for cell in ws[1]:
        cell.fill      = header_fill
        cell.font      = Font(color="000000", bold=True)
        cell.border    = border
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Merge OM Name + alternating colors
    row = 2
    toggle = True
    while row <= ws.max_row:
        om = ws[f"A{row}"].value
        if om:
            start = row
            while row + 1 <= ws.max_row and ws[f"A{row+1}"].value == om:
                row += 1
            end = row
            if end > start:
                ws.merge_cells(start_row=start, start_column=1, end_row=end, end_column=1)
            fill = om_fill1 if toggle else om_fill2
            toggle = not toggle
            for r in range(start, end + 1):
                for c in range(1, 6):
                    cell           = ws.cell(row=r, column=c)
                    cell.fill      = fill
                    cell.border    = border
                    cell.alignment = Alignment(horizontal="center", vertical="center")
        row += 1

    # Auto-fit columns
    for col in ws.columns:
        max_len = max((len(str(c.value)) for c in col if c.value), default=10)
        ws.column_dimensions[col[0].column_letter].width = max_len + 2

    wb.save(output_file)
    print(f"   ✅ Saved: {output_file}")


def process_fleet_file(input_path, output_folder="output", allowed_statuses=None):
    print(f"📊 Loading IC Fleet: {input_path}")
    df = pd.read_excel(input_path, sheet_name=SHEET_NAME)

    # Rename columns to match expected schema
    rename_map = {}
    if "OpsManager" in df.columns:
        rename_map["OpsManager"] = "OM Name"
    if "Client" in df.columns:
        rename_map["Client"] = "Client Name"
    if rename_map:
        df = df.rename(columns=rename_map)

    if allowed_statuses is None:
        from datetime import datetime
        day = datetime.now().day
        if day < 10:
            allowed_statuses = ["committed", "not committed"]
        else:
            allowed_statuses = ["committed", "not committed", "approved", "confirmed"]

    allowed_lower = [s.lower() for s in allowed_statuses]

    base_filtered = df[
        (df["Status"].str.strip().str.lower() == "pay") &
        (df["Site Status"].str.strip().str.lower().isin(allowed_lower)) &
        (df["Site Code"].notna())    & (df["Site Code"].astype(str).str.strip()    != "") &
        (df["OM Name"].notna())      & (df["OM Name"].astype(str).str.strip()      != "") &
        (~df["OM Name"].astype(str).str.strip().str.lower().isin(["dipayan chatterjee", "avik debnath"])) &
        (df["Client Name"].notna())  & (df["Client Name"].astype(str).str.strip()  != "")
    ].copy()

    # Align with output schema by assigning Status values to Payout Status
    base_filtered["Payout Status"] = base_filtered["Status"]

    pivot_df = (
        base_filtered[["OM Name", "Site Code", "Site Status", "Payout Status", "Client Name"]]
        .drop_duplicates(subset=["OM Name", "Site Code"])
        .sort_values(by=["OM Name", "Site Code"])
        .reset_index(drop=True)
    )

    os.makedirs(output_folder, exist_ok=True)
    total_records = len(pivot_df)
    output_file = None
    if total_records > 0:
        output_file = os.path.join(output_folder, "IC_Fleet_Pending.xlsx")
        save_and_style_fleet(pivot_df, output_file)
    else:
        print("   ✅ No IC Fleet pending records.")

    return output_file, total_records
