"""
IC Staff Monthly Payout Processor
===================================
Exact logic from user's script:
  - Sheet: "System Format"
  - Filter: Payout Status = "pay"
  - Filter: Site Status in ["not committed", "committed"]
  - Exclude blank: Site Code, OM Name, Client Name
  - Output columns: OM Name, Site Code, Site Status, Payout Status, Client Name
  - Style: merged OM Name cells, alternating green rows
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
import os

SHEET_NAME          = "System Format"


def save_and_style_staff(pivot_df, output_file):
    pivot_df.to_excel(output_file, sheet_name="STAFF", index=False)

    # ── Style ────────────────────────────────────────────
    wb = load_workbook(output_file)
    ws = wb["STAFF"]

    header_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
    thin        = Side(style="thin")
    border      = Border(left=thin, right=thin, top=thin, bottom=thin)
    om_fill1    = PatternFill(start_color="E6F4EA", end_color="E6F4EA", fill_type="solid")
    om_fill2    = PatternFill(start_color="DFF1DD", end_color="DFF1DD", fill_type="solid")

    # Header
    for cell in ws[1]:
        cell.fill      = header_fill
        cell.font      = Font(color="000000", bold=True)
        cell.border    = border
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Merge OM Name + alternating colors
    row = 2
    toggle = True
    while row <= ws.max_row:
        om = ws[f"A{row}"].value
        if om:
            start = row
            while row + 1 <= ws.max_row and ws[f"A{row+1}"].value == om:
                row += 1
            end = row
            if end > start:
                ws.merge_cells(start_row=start, start_column=1, end_row=end, end_column=1)
            fill = om_fill1 if toggle else om_fill2
            toggle = not toggle
            for r in range(start, end + 1):
                for c in range(1, 6):
                    cell           = ws.cell(row=r, column=c)
                    cell.fill      = fill
                    cell.border    = border
                    cell.alignment = Alignment(horizontal="center", vertical="center")
        row += 1

    # Auto-fit columns
    for col in ws.columns:
        max_len = max((len(str(c.value)) for c in col if c.value), default=10)
        ws.column_dimensions[col[0].column_letter].width = max_len + 2

    wb.save(output_file)
    print(f"   ✅ Saved: {output_file}")


def process_staff_file(input_path, output_folder="output", allowed_statuses=None):
    print(f"📊 Loading IC Staff: {input_path}")
    df = pd.read_excel(input_path, sheet_name=SHEET_NAME)

    if allowed_statuses is None:
        from datetime import datetime
        day = datetime.now().day
        if day < 7:
            allowed_statuses = ["committed", "not committed"]
        else:
            allowed_statuses = ["committed", "not committed", "approved", "confirmed"]

    allowed_lower = [s.lower() for s in allowed_statuses]

    base_filtered = df[
        (df["Payout Status"].str.strip().str.lower() == "pay") &
        (df["Site Code"].notna())    & (df["Site Code"].astype(str).str.strip()    != "") &
        (df["OM Name"].notna())      & (df["OM Name"].astype(str).str.strip()      != "") &
        (df["OM Name"].astype(str).str.strip().str.lower() != "dipayan chatterjee") &
        (df["Client Name"].notna())  & (df["Client Name"].astype(str).str.strip()  != "")
    ]

    # Split 1: Committed (committed, approved, confirmed if allowed)
    comm_list = [s for s in ["committed", "approved", "confirmed"] if s in allowed_lower]
    committed_df = base_filtered[
        base_filtered["Site Status"].str.strip().str.lower().isin(comm_list)
    ]
    pivot_committed = (
        committed_df[["OM Name", "Site Code", "Site Status", "Payout Status", "Client Name"]]
        .drop_duplicates(subset=["OM Name", "Site Code"])
        .sort_values(by=["OM Name", "Site Code"])
        .reset_index(drop=True)
    )

    # Split 2: Not Committed (not committed)
    not_committed_df = base_filtered[
        base_filtered["Site Status"].str.strip().str.lower() == "not committed"
    ]
    pivot_not_committed = (
        not_committed_df[["OM Name", "Site Code", "Site Status", "Payout Status", "Client Name"]]
        .drop_duplicates(subset=["OM Name", "Site Code"])
        .sort_values(by=["OM Name", "Site Code"])
        .reset_index(drop=True)
    )

    os.makedirs(output_folder, exist_ok=True)

    # Process Committed
    total_committed = len(pivot_committed)
    file_committed = None
    if total_committed > 0:
        file_committed = os.path.join(output_folder, "IC_Staff_Committed_Pending.xlsx")
        save_and_style_staff(pivot_committed, file_committed)
    else:
        print("   ✅ No IC Staff Committed pending records.")

    # Process Not Committed
    total_not_committed = len(pivot_not_committed)
    file_not_committed = None
    if total_not_committed > 0:
        file_not_committed = os.path.join(output_folder, "IC_Staff_Not_Committed_Pending.xlsx")
        save_and_style_staff(pivot_not_committed, file_not_committed)
    else:
        print("   ✅ No IC Staff Not Committed pending records.")

    return file_committed, total_committed, file_not_committed, total_not_committed
