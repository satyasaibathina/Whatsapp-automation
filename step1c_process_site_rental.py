"""
Site Rental Payout Processor
==============================
2 files needed:
  1. Site Rental payout Excel (has Status, Site Status, Client, Site Code)
  2. ActiveAndDisabledSites Excel (has Site Code + OM → merged to get OM)

Filter: Status=Pay, Site Status in [not committed, committed]
Output: OM, Site Code, Site Status, Status, Client
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
import os

ALLOWED_SITE_STATUS = ["not committed", "committed"]


def process_site_rental_file(rental_path, active_sites_path, output_folder="output"):
    print(f"   📊 Loading Site Rental: {rental_path}")
    df1 = pd.read_excel(rental_path)

    print(f"   📊 Loading Active Sites: {active_sites_path}")
    df2 = pd.read_excel(active_sites_path)

    # Merge to get OM column from active sites
    df_merged = df1.merge(df2[['Site Code', 'OM']], on='Site Code', how='left')

    # Insert OM right after Site Code
    cols = df_merged.columns.tolist()
    if 'OM' in cols:
        cols.remove('OM')
        site_idx = cols.index('Site Code')
        cols.insert(site_idx + 1, 'OM')
        df_merged = df_merged[cols]

    # Filter
    filtered_df = df_merged[
        (df_merged["Status"].str.strip().str.lower() == "pay") &
        (df_merged["Site Status"].str.strip().str.lower().isin([s.lower() for s in ALLOWED_SITE_STATUS])) &
        (df_merged["OM"].notna())     & (df_merged["OM"].astype(str).str.strip()     != "") &
        (df_merged["Client"].notna()) & (df_merged["Client"].astype(str).str.strip() != "")
    ]

    pivot_df = (
        filtered_df[["OM", "Site Code", "Site Status", "Status", "Client"]]
        .drop_duplicates(subset=["OM", "Site Code"])
        .sort_values(by=["OM", "Site Code"])
        .reset_index(drop=True)
    )

    total = len(pivot_df)
    print(f"   Filtered records: {total}")

    if total == 0:
        print("   ✅ No Site Rental pending records.")
        return None, 0

    os.makedirs(output_folder, exist_ok=True)
    output_file = os.path.join(output_folder, "SiteRental_Pending.xlsx")
    pivot_df.to_excel(output_file, sheet_name="SiteRental", index=False)

    wb = load_workbook(output_file)
    ws = wb["SiteRental"]

    header_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
    thin        = Side(style="thin")
    border      = Border(left=thin, right=thin, top=thin, bottom=thin)
    om_fill1    = PatternFill(start_color="E6F4EA", end_color="E6F4EA", fill_type="solid")
    om_fill2    = PatternFill(start_color="DFF1DD", end_color="DFF1DD", fill_type="solid")

    for cell in ws[1]:
        cell.fill      = header_fill
        cell.font      = Font(color="000000", bold=True)
        cell.border    = border
        cell.alignment = Alignment(horizontal="center", vertical="center")

    row = 2
    toggle = True
    while row <= ws.max_row:
        om = ws[f"A{row}"].value
        if om:
            start = row
            while row + 1 <= ws.max_row and ws[f"A{row+1}"].value == om:
                row += 1
            end = row
            if end > start:
                ws.merge_cells(start_row=start, start_column=1, end_row=end, end_column=1)
            fill = om_fill1 if toggle else om_fill2
            toggle = not toggle
            for r in range(start, end + 1):
                for c in range(1, 6):
                    cell           = ws.cell(row=r, column=c)
                    cell.fill      = fill
                    cell.border    = border
                    cell.alignment = Alignment(horizontal="center", vertical="center")
        row += 1

    for col in ws.columns:
        max_len = max((len(str(c.value)) for c in col if c.value), default=10)
        ws.column_dimensions[col[0].column_letter].width = max_len + 2

    wb.save(output_file)
    print(f"   ✅ Saved: {output_file}")
    return output_file, total
