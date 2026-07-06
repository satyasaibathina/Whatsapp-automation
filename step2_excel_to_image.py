"""
STEP 2 — Excel → PNG Image
"""
import os
from openpyxl import load_workbook
from PIL import Image, ImageDraw, ImageFont

FONT_SIZE  = 14
ROW_HEIGHT = 36
SCALE      = 4


def load_font(bold=False):
    candidates = (
        ["/usr/share/fonts/truetype/liberation/LiberationSans-Bold.ttf",
         "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf",
         "C:\\Windows\\Fonts\\arialbd.ttf"]
        if bold else
        ["/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf",
         "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
         "C:\\Windows\\Fonts\\arial.ttf"]
    )
    for p in candidates:
        if os.path.exists(p):
            try: return ImageFont.truetype(p, FONT_SIZE * SCALE)
            except: pass
    return ImageFont.load_default()


def hex_to_rgb(h):
    h = h.lstrip("#")
    if len(h) == 8: h = h[2:]
    if len(h) == 6: return tuple(int(h[i:i+2], 16) for i in (0, 2, 4))
    return (255, 255, 255)


def get_bg(cell):
    f = cell.fill
    if f and f.fill_type == "solid" and f.fgColor and f.fgColor.type == "rgb":
        return hex_to_rgb(f.fgColor.rgb)
    return (255, 255, 255)


def excel_to_image(excel_path, output_path=None):
    print(f"🖼️  Converting: {excel_path}")
    wb = load_workbook(excel_path)
    ws = wb[wb.sheetnames[0]]

    freg  = load_font(False)
    fbold = load_font(True)

    letters = [ws.cell(1, i+1).column_letter for i in range(ws.max_column)]
    widths  = {}
    for col in ws.columns:
        l = col[0].column_letter
        dim = ws.column_dimensions.get(l)
        cw  = dim.width if (dim and dim.width and dim.width > 1) else 12
        widths[l] = max(int(cw * FONT_SIZE * 0.72), 80) * SCALE

    merged = {}
    for mc in ws.merged_cells.ranges:
        for r in range(mc.min_row, mc.max_row+1):
            for c in range(mc.min_col, mc.max_col+1):
                merged[(r,c)] = (mc.min_row, mc.min_col, mc.max_row, mc.max_col)

    W = sum(widths.get(letters[c], 120*SCALE) for c in range(ws.max_column)) + 20*SCALE
    H = (ws.max_row + 1) * ROW_HEIGHT * SCALE + 20*SCALE

    img  = Image.new("RGB", (W, H), (255, 255, 255))
    draw = ImageDraw.Draw(img)

    y = 10*SCALE
    for ri in range(1, ws.max_row+1):
        x = 10*SCALE
        for ci in range(1, ws.max_column+1):
            l    = ws.cell(1, ci).column_letter
            cw   = widths.get(l, 120*SCALE)
            cell = ws.cell(ri, ci)
            mc   = merged.get((ri, ci))

            if mc:
                mr, mc2, mr2, mc3 = mc
                if ri != mr or ci != mc2:
                    x += cw; continue
                sh = (mr2-mr+1) * ROW_HEIGHT*SCALE
                sw = sum(widths.get(ws.cell(1,c).column_letter, 120*SCALE) for c in range(mc2, mc3+1))
            else:
                sh, sw = ROW_HEIGHT*SCALE, cw

            draw.rectangle([x, y, x+sw-1, y+sh-1], fill=get_bg(cell))
            draw.rectangle([x, y, x+sw-1, y+sh-1], outline=(180,180,180), width=SCALE)

            if cell.value:
                txt  = str(cell.value)
                fnt  = fbold if (cell.font and cell.font.bold) else freg
                tc   = (0,0,0)
                if cell.font and cell.font.color and cell.font.color.type=="rgb":
                    tc = hex_to_rgb(cell.font.color.rgb)
                bb   = draw.textbbox((0,0), txt, font=fnt)
                tw,th = bb[2]-bb[0], bb[3]-bb[1]
                draw.text((x+(sw-tw)//2, y+(sh-th)//2), txt, fill=tc, font=fnt)
            x += cw
        y += ROW_HEIGHT*SCALE

    if not output_path:
        output_path = os.path.splitext(excel_path)[0] + ".png"
    img.save(output_path, "PNG", dpi=(300, 300))
    print(f"✅ Image: {output_path}")
    return output_path
